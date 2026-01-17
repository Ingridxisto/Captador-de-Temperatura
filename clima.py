import customtkinter as ctk
import requests
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, PatternFill
from datetime import datetime
from dotenv import load_dotenv
import os

load_dotenv()
API_KEY = os.getenv("API_KEY")
ARQUIVO_EXCEL = "dados_climaticos.xlsx"


def buscar_clima():
    cidade = entrada_cidade.get().strip()

    if not cidade:
        cidade = "São Paulo"

    if not API_KEY:
        label_resultado.configure(text="❌ API_KEY não configurada no .env.")
        return

    url = (
        f"https://api.openweathermap.org/data/2.5/weather"
        f"?q={cidade}&appid={API_KEY}&units=metric&lang=pt_br"
    )

    resposta = requests.get(url)
    dados = resposta.json()

    if resposta.status_code != 200:
        label_resultado.configure(text="❌ Cidade não encontrada.")
        return

    temperatura = dados["main"]["temp"]
    umidade = dados["main"]["humidity"]
    ceu = dados["weather"][0]["description"].capitalize()

    label_resultado.configure(
        text=(
            f"📍 Cidade: {cidade}\n"
            f"🌡 Temperatura: {temperatura} °C\n"
            f"💧 Umidade: {umidade}%\n"
            f"☁ Céu: {ceu}"
        )
    )

    salvar_dados(cidade, temperatura, umidade, ceu)


def salvar_dados(cidade, temperatura, umidade, ceu):
    if not os.path.exists(ARQUIVO_EXCEL):
        arquivo = Workbook()
        planilha = arquivo.active
        planilha.title = "Clima"

        cabecalhos = ["Data e Hora", "Cidade", "Temperatura (°C)", "Umidade (%)", "Condição do Céu"]
        for col, cabecalho in enumerate(cabecalhos, start=1):
            cell = planilha.cell(row=1, column=col)
            cell.value = cabecalho
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            cell.fill = PatternFill(start_color="FFD700", fill_type="solid")
    else:
        arquivo = load_workbook(ARQUIVO_EXCEL)
        planilha = arquivo["Clima"]

    linha = planilha.max_row + 1
    data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    valores = [data_hora, cidade, temperatura, umidade, ceu]

    for col, valor in enumerate(valores, start=1):
        cell = planilha.cell(row=linha, column=col)
        cell.value = valor
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    arquivo.save(ARQUIVO_EXCEL)


# ================= INTERFACE =================
ctk.set_appearance_mode("system")
ctk.set_default_color_theme("blue")
root = ctk.CTk()
root.title("Consulta de Clima")
root.geometry("350x260")

ctk.CTkLabel(root, text="Digite a cidade:", font=("Arial", 10)).pack(pady=5)
entrada_cidade = ctk.CTkEntry(root, width=220)
entrada_cidade.pack(pady=5)

ctk.CTkButton(root, text="Buscar Clima", command=buscar_clima, width=200).pack(pady=10)

label_resultado = ctk.CTkLabel(root, text="", font=("Arial", 10))
label_resultado.pack(pady=10)

root.mainloop()
